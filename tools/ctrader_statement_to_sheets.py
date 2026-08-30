#!/usr/bin/env python3
"""Convert a cTrader "End of Day Statement" HTML file into Google Sheets ready files.

The statement produced by cTrader is built for printing: nested tables, totals rows
written in white-on-white, ``<nobr>`` wrappers, dd/mm/yyyy timestamps and numbers
stored as text. Pasted into Google Sheets nothing is typed correctly and the totals
rows are invisible.

This script parses the statement and writes:

* ``--xlsx``     a multi sheet workbook (real dates, real numbers, formulas,
                 derived trade metrics). Upload it to Drive and Sheets opens it
                 natively, with no import dialog and no locale guessing.
* ``--csv-dir``  one CSV per section, in two flavours:
                 ``*.csv``    comma separated, dot decimals, ISO timestamps
                 ``*.it.csv`` semicolon separated, comma decimals, dd/mm/yyyy
                 (for spreadsheets whose locale is Italian).

Only the standard library is required for CSV output; ``openpyxl`` is needed for
``--xlsx``.

Usage:
    python3 tools/ctrader_statement_to_sheets.py STATEMENT.html \
        --xlsx out/statement.xlsx --csv-dir out/csv
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import math
import os
import re
import sys
from html.parser import HTMLParser

# --------------------------------------------------------------------------- #
# HTML parsing
# --------------------------------------------------------------------------- #

TITLES = ("History", "Positions", "Orders", "Transactions", "Summary")

# Italian labels used for the generated sheets / files.
SECTION_LABELS = {
    "History": "Operazioni chiuse",
    "Positions": "Posizioni aperte",
    "Orders": "Ordini pendenti",
    "Transactions": "Transazioni",
    "Summary": "Riepilogo",
}

HEADER_LABELS = {
    "ID": "ID",
    "Symbol": "Simbolo",
    "Opening Direction": "Direzione apertura",
    "Closing Direction": "Direzione chiusura",
    "Opening Time (UTC+0)": "Apertura (UTC)",
    "Closing Time (UTC+0)": "Chiusura (UTC)",
    "Created (UTC+0)": "Creata (UTC)",
    "Submitted Time (UTC+0)": "Inviato (UTC)",
    "Time (UTC+0)": "Data/ora (UTC)",
    "Entry Price": "Prezzo entrata",
    "Closing Price": "Prezzo uscita",
    "Submitted Price": "Prezzo ordine",
    "Closing Quantity": "Quantita",
    "Current Quantity": "Quantita",
    "Quantity": "Quantita",
    "Current Volume": "Volume",
    "Volume": "Volume",
    "Direction": "Direzione",
    "Order Type": "Tipo ordine",
    "Swap": "Swap",
    "Commission": "Commissioni",
    "Commissions": "Commissioni",
    "Conversion Rate": "Tasso conversione",
    "S/L": "Stop loss",
    "SL is guaranteed": "SL garantito",
    "T/P": "Take profit",
    "TIF": "Validita",
    "Expiry Time": "Scadenza",
    "Type": "Tipo",
    "Note": "Nota",
    "Gross EUR": "Lordo EUR",
    "Net EUR": "Netto EUR",
    "Balance EUR": "Saldo EUR",
    "Amount EUR": "Importo EUR",
}

SUMMARY_LABELS = {
    "Deposit": "Depositi",
    "Withdrawal": "Prelievi",
    "Total Net": "Movimenti netti",
    "Balance": "Saldo",
    "Unrealized P&L": "P/L non realizzato",
    "Realized P&L": "P/L realizzato",
    "Margin": "Margine utilizzato",
    "Free Margin": "Margine libero",
    "Margin Level": "Livello margine",
    "Active bonus": "Bonus attivo",
    "Equity": "Equity",
}

META_LABELS = {
    "Account": "Conto",
    "Account type": "Tipo conto",
    "Currency": "Valuta",
}

GENERATED_RE = re.compile(r"^(\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2}(?:\.\d+)?),?\s*UTC$")
PERIOD_RE = re.compile(r"(\d{2}/\d{2}/\d{4})\s*-\s*(\d{2}/\d{2}/\d{4})")


class StatementParser(HTMLParser):
    """Collect every table of the statement as a list of rows of cell text."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.tables: list[dict] = []
        self._stack: list[dict] = []
        self._cell: list[str] | None = None

    # -- helpers ----------------------------------------------------------- #
    @property
    def _current(self) -> dict | None:
        return self._stack[-1] if self._stack else None

    def handle_starttag(self, tag, attrs):
        attrs = dict(attrs)
        if tag == "table":
            self._stack.append({"class": attrs.get("class", ""), "rows": []})
        elif tag == "tr" and self._current is not None:
            self._current["rows"].append([])
        elif tag in ("td", "th") and self._current is not None:
            if not self._current["rows"]:
                self._current["rows"].append([])
            self._cell = []

    def handle_endtag(self, tag):
        if tag in ("td", "th") and self._cell is not None and self._current is not None:
            text = re.sub(r"\s+", " ", "".join(self._cell)).strip()
            self._current["rows"][-1].append(text)
            self._cell = None
        elif tag == "table" and self._stack:
            self.tables.append(self._stack.pop())

    def handle_data(self, data):
        if self._cell is not None:
            self._cell.append(data)


def read_html(path: str) -> str:
    """cTrader declares charset=utf-16 but ships utf-8/ascii; try both."""
    raw = open(path, "rb").read()
    for encoding in ("utf-8-sig", "utf-8", "utf-16", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding)
        except (UnicodeDecodeError, UnicodeError):
            continue
        if "<html" in text.lower():
            return text
    return raw.decode("utf-8", "replace")


# --------------------------------------------------------------------------- #
# Value typing
# --------------------------------------------------------------------------- #

DATE_RE = re.compile(r"^(\d{2})/(\d{2})/(\d{4})(?:[ T](\d{2}):(\d{2}):(\d{2})(?:\.(\d+))?)?$")
NUM_RE = re.compile(r"^-?\d+(\.\d+)?$")


def typed(value: str):
    """Return a datetime / float / str for a raw statement cell."""
    value = (value or "").strip()
    if value in ("", "-", "n/a", "N/A"):
        return None if value in ("", "-") else value
    m = DATE_RE.match(value)
    if m:
        day, month, year, hh, mm, ss, ms = m.groups()
        if hh is None:
            return dt.date(int(year), int(month), int(day))
        micro = int((ms or "0").ljust(6, "0")[:6])
        return dt.datetime(int(year), int(month), int(day), int(hh), int(mm), int(ss), micro)
    if NUM_RE.match(value):
        return float(value)
    return value


def parse_statement(path: str) -> dict:
    parser = StatementParser()
    parser.feed(read_html(path))

    meta: dict[str, str] = {}
    sections: dict[str, dict] = {}
    summary: list[tuple[str, object]] = []
    period = generated = None

    for table in parser.tables:
        rows = [r for r in table["rows"] if any(c for c in r)]
        if not rows:
            continue
        cls = table["class"]

        if cls == "dataTable":
            title = rows[0][0] if rows[0] else ""
            if title not in TITLES:
                continue
            header = [c for c in rows[1][1:]] if len(rows) > 1 else []
            body, totals = [], []
            for row in rows[2:]:
                cells = row[1:]
                if row and row[0] == "Totals":
                    totals = cells
                elif len(cells) >= max(1, len(header) - 2) and "No " not in row[0]:
                    body.append([typed(c) for c in cells])
            sections[title] = {"headers": header, "rows": body, "totals": totals}

        elif cls == "summaryTable":
            for row in rows[1:]:
                cells = [c for c in row if c != ""]
                for i in range(0, len(cells) - 1, 2):
                    label, value = cells[i], cells[i + 1]
                    if label and not label.startswith("Summary"):
                        summary.append((SUMMARY_LABELS.get(label, label), typed(value)))

        else:  # header / meta tables
            for row in rows:
                cells = [c for c in row if c]
                for i, text in enumerate(cells):
                    label = text.rstrip(" :")
                    if label in META_LABELS and i + 1 < len(cells):
                        meta[META_LABELS[label]] = cells[i + 1]
                    elif GENERATED_RE.match(text):
                        generated = GENERATED_RE.match(text).group(1)
                    elif PERIOD_RE.search(text):
                        m2 = PERIOD_RE.search(text)
                        period = f"{m2.group(1)} - {m2.group(2)}"
                    elif "@" in text and "." in text and " " not in text:
                        meta["Email"] = text
                    elif (len(cells) == 1 and "Email" in meta and "Intestatario" not in meta
                          and len(re.findall(r"[A-Za-z]{2,}", text)) >= 2):
                        meta["Intestatario"] = text.strip()
    if period:
        meta["Periodo"] = period
    if generated:
        meta["Generato"] = f"{generated} UTC"
    return {"meta": meta, "sections": sections, "summary": summary}


# --------------------------------------------------------------------------- #
# Derived trade metrics and column layout
# --------------------------------------------------------------------------- #

def round_half_up(value: float, digits: int) -> float:
    """Round half away from zero, the way a spreadsheet (and the page) does.

    Python's built-in round() goes to the nearest even digit, so a duration of
    330.445 minutes becomes 330.44 here and 330.45 everywhere else.
    """
    factor = 10 ** digits
    rounded = math.floor(abs(value) * factor + 0.5) / factor
    return rounded if value >= 0 else -rounded


def derive_history(headers: list[str], row: list) -> dict:
    """Duration, price move in points, win/loss flag and % of balance."""
    idx = {h: i for i, h in enumerate(headers)}

    def get(name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    opened, closed = get("Opening Time (UTC+0)"), get("Closing Time (UTC+0)")
    duration = None
    if isinstance(opened, dt.datetime) and isinstance(closed, dt.datetime):
        duration = round_half_up((closed - opened).total_seconds() / 60, 2)

    entry, exit_ = get("Entry Price"), get("Closing Price")
    direction = get("Opening Direction")
    points = None
    if isinstance(entry, float) and isinstance(exit_, float) and direction:
        points = round_half_up(exit_ - entry if direction.upper() == "BUY" else entry - exit_, 5)

    net, balance = get("Net EUR"), get("Balance EUR")
    outcome = pct = None
    if isinstance(net, float):
        outcome = "WIN" if net > 0 else ("LOSS" if net < 0 else "BE")
        if isinstance(balance, float) and balance - net:
            pct = net / (balance - net)
    return {"Durata (min)": duration, "Punti": points,
            "Esito": outcome, "P/L % sul saldo": pct}


# Output columns, in reading order: identity, timing, size and prices, result,
# running balance. Entries are (label, source) where a source not present in the
# statement is simply skipped, and "@" marks a derived column.
VIEWS = {
    "History": [
        ("ID", "ID"), ("Simbolo", "Symbol"), ("Direzione", "Opening Direction"),
        ("Apertura (UTC)", "Opening Time (UTC+0)"), ("Chiusura (UTC)", "Closing Time (UTC+0)"),
        ("Durata (min)", "@Durata (min)"), ("Quantita", "Closing Quantity"),
        ("Prezzo entrata", "Entry Price"), ("Prezzo uscita", "Closing Price"),
        ("Punti", "@Punti"), ("Lordo EUR", "Gross EUR"), ("Swap", "Swap"),
        ("Commissioni", "Commission"), ("Netto EUR", "Net EUR"), ("Esito", "@Esito"),
        ("P/L % sul saldo", "@P/L % sul saldo"), ("Saldo EUR", "Balance EUR"),
        ("Tasso conversione", "Conversion Rate"),
    ],
    "Orders": [
        ("ID", "ID"), ("Simbolo", "Symbol"), ("Direzione", "Direction"),
        ("Tipo", "Order Type"), ("Inviato (UTC)", "Submitted Time (UTC+0)"),
        ("Prezzo ordine", "Submitted Price"), ("Quantita", "Current Quantity"),
        ("Volume", "Current Volume"), ("Stop loss", "S/L"), ("Take profit", "T/P"),
        ("Validita", "TIF"), ("Scadenza", "Expiry Time"), ("SL garantito", "SL is guaranteed"),
    ],
    "Positions": [
        ("ID", "ID"), ("Simbolo", "Symbol"), ("Direzione", "Direction"),
        ("Creata (UTC)", "Created (UTC+0)"), ("Prezzo entrata", "Entry Price"),
        ("Quantita", "Quantity"), ("Volume", "Volume"), ("Stop loss", "S/L"),
        ("Take profit", "T/P"), ("Swap", "Swap"), ("Commissioni", "Commissions"),
        ("Lordo EUR", "Gross EUR"), ("Netto EUR", "Net EUR"),
        ("SL garantito", "SL is guaranteed"),
    ],
    "Transactions": [
        ("ID", "ID"), ("Data/ora (UTC)", "Time (UTC+0)"), ("Tipo", "Type"),
        ("Importo EUR", "Amount EUR"), ("Nota", "Note"),
    ],
}

TOTAL_COLUMNS = ("Lordo EUR", "Swap", "Commissioni", "Netto EUR")

# Per-instrument aggregates, keyed by the symbol sitting in column A of the same
# row. {row} is the spreadsheet row the cell lands on, ARG the locale's argument
# separator.
SYMBOL_AGGREGATES = [
    ("Operazioni", "count", "=COUNTIF({sym}ARG$A{row})"),
    ("Vinte", "wins", '=COUNTIFS({sym}ARG$A{row}ARG{net}ARG">0")'),
    ("Perse", "losses", '=COUNTIFS({sym}ARG$A{row}ARG{net}ARG"<0")'),
    ("Win rate", "winrate",
     '=IFERROR(COUNTIFS({sym}ARG$A{row}ARG{net}ARG">0")/COUNTIF({sym}ARG$A{row})ARG"")'),
    ("Netto EUR", "net", "=SUMIF({sym}ARG$A{row}ARG{net})"),
    ("Netto medio", "avg", '=IFERROR(AVERAGEIF({sym}ARG$A{row}ARG{net})ARG"")'),
]


def view(name: str, section: dict) -> tuple[list[str], list[list]]:
    """Re-order a parsed section into its output columns."""
    spec = VIEWS.get(name)
    if not spec:
        labels = [HEADER_LABELS.get(h, h) for h in section["headers"]]
        return labels, [list(r) for r in section["rows"]]

    idx = {h: i for i, h in enumerate(section["headers"])}
    columns = [(label, source) for label, source in spec
               if source.startswith("@") or source in idx]
    rows = []
    for row in section["rows"]:
        derived = derive_history(section["headers"], row) if name == "History" else {}
        rows.append([derived.get(source[1:]) if source.startswith("@")
                     else (row[idx[source]] if idx[source] < len(row) else None)
                     for _, source in columns])
    return [label for label, _ in columns], rows


# --------------------------------------------------------------------------- #
# Statistics
# --------------------------------------------------------------------------- #

# Each entry is (label, statistics key, formula template). The template is
# rendered by whichever writer knows where the trades table ended up; ranges are
# substituted for {net}, {dur} and {sym}, and ARG is the locale's argument
# separator (';' for Italian spreadsheets).
STATS = [
    ("Operazioni chiuse", "count", "=COUNT({net})"),
    ("Operazioni in profitto", "wins", '=COUNTIF({net}ARG">0")'),
    ("Operazioni in perdita", "losses", '=COUNTIF({net}ARG"<0")'),
    ("Win rate", "win_rate", '=IFERROR(COUNTIF({net}ARG">0")/COUNT({net})ARG"")'),
    ("Profitto medio", "avg_win", '=IFERROR(AVERAGEIF({net}ARG">0")ARG"")'),
    ("Perdita media", "avg_loss", '=IFERROR(AVERAGEIF({net}ARG"<0")ARG"")'),
    ("Profit factor", "profit_factor",
     '=IFERROR(SUMIF({net}ARG">0")/ABS(SUMIF({net}ARG"<0"))ARG"")'),
    ("Aspettativa per operazione", "expectancy", '=IFERROR(AVERAGE({net})ARG"")'),
    ("Migliore operazione", "best", "=MAX({net})"),
    ("Peggiore operazione", "worst", "=MIN({net})"),
    ("Durata media (min)", "avg_duration", '=IFERROR(AVERAGE({dur})ARG"")'),
    ("Strumenti negoziati", "symbols",
     '=IFERROR(SUMPRODUCT((COUNTIF({sym}ARG{sym})>0)/COUNTIF({sym}ARG{sym}))ARG"")'),
]


def compute_stats(data: dict) -> dict:
    history = data["sections"].get("History")
    if not history or not history["rows"]:
        return {}
    idx = history["headers"].index("Net EUR")
    nets = [r[idx] for r in history["rows"] if isinstance(r[idx], float)]
    wins = [v for v in nets if v > 0]
    losses = [v for v in nets if v < 0]
    durations = [derive_history(history["headers"], r)["Durata (min)"]
                 for r in history["rows"]]
    durations = [d for d in durations if d is not None]
    symbols = {r[history["headers"].index("Symbol")] for r in history["rows"]}
    return {
        "count": len(nets), "wins": len(wins), "losses": len(losses),
        "win_rate": len(wins) / len(nets) if nets else None,
        "total": round(sum(nets), 10),
        "avg_win": round(sum(wins) / len(wins), 10) if wins else None,
        "avg_loss": round(sum(losses) / len(losses), 10) if losses else None,
        "profit_factor": round(sum(wins) / abs(sum(losses)), 10) if losses else None,
        "expectancy": round(sum(nets) / len(nets), 10) if nets else None,
        "best": max(nets) if nets else None, "worst": min(nets) if nets else None,
        "avg_duration": round(sum(durations) / len(durations), 10) if durations else None,
        "symbols": len(symbols),
    }


def by_symbol(data: dict) -> list[list]:
    """Per-instrument tally, worst impact first: [simbolo, operazioni, vinte, perse, netto]."""
    history = data["sections"].get("History")
    if not history or not history["rows"]:
        return []
    headers, rows = view("History", history)
    i_sym, i_net = headers.index("Simbolo"), headers.index("Netto EUR")
    tally: dict[str, list] = {}
    for row in rows:
        entry = tally.setdefault(row[i_sym], [0, 0, 0, 0.0])
        entry[0] += 1
        net = row[i_net]
        if isinstance(net, float):
            if net > 0:
                entry[1] += 1
            elif net < 0:
                entry[2] += 1
            entry[3] += net
    ordered = sorted(tally.items(), key=lambda kv: -abs(kv[1][3]))
    return [[symbol] + counts[:3] + [round(counts[3], 10)] for symbol, counts in ordered]


def by_day(data: dict) -> list[list]:
    """Trading days in order: [data, operazioni, netto, saldo di fine giornata]."""
    history = data["sections"].get("History")
    if not history or not history["rows"]:
        return []
    headers, rows = view("History", history)
    i_close = headers.index("Chiusura (UTC)")
    i_net, i_balance = headers.index("Netto EUR"), headers.index("Saldo EUR")
    days: dict[dt.date, list] = {}
    for row in rows:
        closed = row[i_close]
        if not isinstance(closed, dt.datetime):
            continue
        entry = days.setdefault(closed.date(), [0, 0.0, None, None])
        entry[0] += 1
        if isinstance(row[i_net], float):
            entry[1] += row[i_net]
        if entry[2] is None or closed >= entry[2]:      # last close of the day sets the balance
            entry[2], entry[3] = closed, row[i_balance]
    return [[day, n, round(net, 10), balance]
            for day, (n, net, _, balance) in sorted(days.items())]


# Fixed bins, coarse to keep the histogram readable: scalps, minutes, hours.
DURATION_BUCKETS = [("meno di 1 min", 0, 1), ("1-5 min", 1, 5), ("5-15 min", 5, 15),
                    ("15-60 min", 15, 60), ("1-4 ore", 60, 240),
                    ("oltre 4 ore", 240, float("inf"))]


def duration_buckets(data: dict) -> list[list]:
    """How long trades were held: [fascia, operazioni, netto]."""
    history = data["sections"].get("History")
    if not history or not history["rows"]:
        return []
    headers, rows = view("History", history)
    i_dur, i_net = headers.index("Durata (min)"), headers.index("Netto EUR")
    tally = {label: [0, 0.0] for label, _, _ in DURATION_BUCKETS}
    for row in rows:
        duration = row[i_dur]
        if not isinstance(duration, float):
            continue
        for label, low, high in DURATION_BUCKETS:
            if low <= duration < high:
                tally[label][0] += 1
                if isinstance(row[i_net], float):
                    tally[label][1] += row[i_net]
                break
    return [[label, count, round(net, 10)]
            for label, (count, net) in tally.items() if count]


# --------------------------------------------------------------------------- #
# Report layout
# --------------------------------------------------------------------------- #

def summary_value(data: dict, *labels):
    for label, value in data["summary"]:
        if label in labels:
            return value
    return None


def build_report(data: dict) -> list[dict]:
    """The statement as ordered blocks, most decision-relevant first.

    Blocks carry a group so a writer can either stack them (one CSV, one Google
    Sheet tab) or split them across sheets. Formula cells are emitted as tokens
    ({{SUM:label}}, {{STAT:key}}) and resolved once the writer knows the row
    numbers the trades table landed on.
    """
    meta = data["meta"]
    blocks: list[dict] = []
    get = lambda *labels: summary_value(data, *labels)

    subtitle = " · ".join(x for x in [
        f"Conto {meta.get('Conto', '-')}", meta.get("Tipo conto"),
        meta.get("Valuta"), meta.get("Periodo")] if x)
    blocks.append({"group": "summary", "kind": "heading",
                   "title": "ESTRATTO CONTO cTrader", "subtitle": subtitle})

    blocks.append({"group": "summary", "kind": "pairs", "title": "SITUAZIONE DEL CONTO",
                   "rows": [[label, get(*names)] for label, *names in [
                       ["Saldo", "Saldo", "Balance"],
                       ["Equity", "Equity"],
                       ["Margine libero", "Margine libero", "Free Margin"],
                       ["Margine utilizzato", "Margine utilizzato", "Margin"],
                       ["Livello margine", "Livello margine", "Margin Level"],
                       ["P/L non realizzato", "P/L non realizzato", "Unrealized P&L"],
                       ["Bonus attivo", "Bonus attivo", "Active bonus"],
                   ]]})

    stats = compute_stats(data)
    performance = [["P/L realizzato", get("P/L realizzato", "Realized P&L")]]
    if stats:
        performance += [[label, "{{STAT:%s}}" % key] for label, key, _ in STATS]
    performance += [
        ["Depositi", get("Depositi", "Deposit")],
        ["Prelievi", get("Prelievi", "Withdrawal")],
        ["Movimenti netti", get("Movimenti netti", "Total Net")],
    ]
    blocks.append({"group": "summary", "kind": "pairs",
                   "title": "PERFORMANCE DEL PERIODO", "rows": performance})

    symbols = by_symbol(data)
    if symbols:
        blocks.append({
            "group": "summary", "id": "symbols",
            "kind": "table", "title": "RIEPILOGO PER STRUMENTO",
            "headers": ["Simbolo"] + [label for label, _, _ in SYMBOL_AGGREGATES],
            "rows": [[row[0]] + ["{{SYMAGG:%s:%s}}" % (key, row[0])
                                 for _, key, _ in SYMBOL_AGGREGATES] for row in symbols],
            "totals": None})

    durations = duration_buckets(data)
    if durations:
        blocks.append({
            "group": "summary", "id": "durations",
            "kind": "table", "title": "DISTRIBUZIONE DURATE",
            "headers": ["Fascia", "Operazioni", "Netto EUR"],
            "rows": durations, "totals": None})

    days = by_day(data)
    if days:
        blocks.append({
            "group": "daily", "id": "daily",
            "kind": "table", "title": "ANDAMENTO GIORNALIERO",
            "headers": ["Data", "Operazioni", "Netto EUR", "Saldo fine giornata"],
            "rows": days, "totals": None, "is_daily": True})

    for name, group in (("History", "trades"), ("Orders", "orders"),
                        ("Positions", "positions"), ("Transactions", "transactions")):
        section = data["sections"].get(name)
        if not section:
            continue
        headers, rows = view(name, section)
        block = {"group": group, "kind": "table", "title": SECTION_LABELS[name].upper(),
                 "headers": headers, "rows": rows, "totals": None,
                 "is_trades": name == "History"}
        if name == "History" and rows:
            totals = [""] * len(headers)
            totals[0] = "TOTALE"
            for label in TOTAL_COLUMNS:
                if label in headers:
                    totals[headers.index(label)] = "{{SUM:%s}}" % label
            block["totals"] = totals
        blocks.append(block)

    blocks.append({"group": "summary", "kind": "pairs", "title": "DATI DEL CONTO",
                   "rows": [[k, v] for k, v in meta.items()]})
    return blocks


def resolve(cell, ctx):
    """Turn a {{SUM:...}} / {{STAT:...}} token into a formula for one writer."""
    if not isinstance(cell, str) or not cell.startswith("{{"):
        return cell
    kind, _, argument = cell[2:-2].partition(":")
    rng = lambda col: f"{ctx['sheet']}${col}${ctx['first']}:${col}${ctx['last']}"
    if kind == "SUM":
        return f"=SUM({rng(ctx['column'](argument))})"
    if kind == "SYMAGG":
        argument = argument.split(":")[0]           # the symbol rides along for value output
        template = next(t for _, key, t in SYMBOL_AGGREGATES if key == argument)
        return (template.replace("ARG", ctx["arg"])
                        .format(sym=rng(ctx["column"]("Simbolo")),
                                net=rng(ctx["column"]("Netto EUR")),
                                row=ctx["row"]))
    template = next(t for label, key, t in STATS if key == argument)
    ranges = {}
    for token, label in (("net", "Netto EUR"), ("dur", "Durata (min)"), ("sym", "Simbolo")):
        col = ctx["column"](label)
        ranges[token] = rng(col) if col else ""
    return template.replace("ARG", ctx["arg"]).format(**ranges)


def check_consistency(data: dict) -> list[str]:
    """Cross-check the parsed rows against the totals printed on the statement."""
    warnings = []
    history = data["sections"].get("History")
    if not history or not history["rows"]:
        return warnings
    idx = {h: i for i, h in enumerate(history["headers"])}
    net = sum(r[idx["Net EUR"]] for r in history["rows"]
              if isinstance(r[idx["Net EUR"]], float))
    for label, value in data["summary"]:
        if label in ("P/L realizzato", "Realized P&L") and isinstance(value, float):
            if abs(net - value) > 0.01:
                warnings.append(f"somma Netto EUR ({net:.2f}) != P/L realizzato dello "
                                f"statement ({value:.2f})")
    last_balance = history["rows"][-1][idx["Balance EUR"]]
    for label, value in data["summary"]:
        if label in ("Saldo", "Balance") and isinstance(value, float):
            if isinstance(last_balance, float) and abs(last_balance - value) > 0.01:
                warnings.append(f"ultimo Saldo EUR ({last_balance:.2f}) != saldo di "
                                f"riepilogo ({value:.2f})")
    return warnings


# --------------------------------------------------------------------------- #
# Formatting helpers
# --------------------------------------------------------------------------- #

def plain(value: float) -> str:
    """Decimal notation always: a spreadsheet reads 8,4834e-06 as text, not a number."""
    text = repr(round(value, 10))
    if "e" in text or "E" in text:
        text = f"{value:.10f}".rstrip("0").rstrip(".") or "0"
    return text


def fmt_intl(value):
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        return plain(value)
    return "" if value is None else value


def fmt_it(value):
    if isinstance(value, dt.datetime):
        return value.strftime("%d/%m/%Y %H.%M.%S")
    if isinstance(value, dt.date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, float):
        return plain(value).replace(".", ",")
    return "" if value is None else value


def col_letter(index: int) -> str:
    letters = ""
    while index:
        index, rest = divmod(index - 1, 26)
        letters = chr(65 + rest) + letters
    return letters


# --------------------------------------------------------------------------- #
# CSV output
# --------------------------------------------------------------------------- #

def write_csvs(data: dict, out_dir: str) -> list[str]:
    """One CSV per section, in both dot-decimal and Italian-locale flavours."""
    os.makedirs(out_dir, exist_ok=True)
    written = []
    blocks: list[tuple[str, list[str], list[list]]] = [
        ("00_dati_conto", ["Voce", "Valore"], [[k, v] for k, v in data["meta"].items()])
    ]
    for i, name in enumerate(TITLES[:-1], start=1):
        section = data["sections"].get(name)
        if not section:
            continue
        headers, rows = view(name, section)
        blocks.append((f"{i:02d}_{SECTION_LABELS[name].replace(' ', '_').lower()}",
                       headers, rows))
    blocks.append(("05_riepilogo", ["Voce", "Valore"],
                   [[k, v] for k, v in data["summary"]]))

    for base, headers, rows in blocks:
        for suffix, delim, fmt in ((".csv", ",", fmt_intl), (".it.csv", ";", fmt_it)):
            path = os.path.join(out_dir, base + suffix)
            with open(path, "w", newline="", encoding="utf-8") as fh:
                writer = csv.writer(fh, delimiter=delim)
                writer.writerow(headers)
                writer.writerows([[fmt(c) for c in row] for row in rows])
            written.append(path)
    return written


def write_gsheet_csv(data: dict, path: str) -> str:
    """One semicolon/comma-decimal CSV holding every block, stacked.

    Google Drive converts an uploaded CSV into a native Google Sheet, but only
    ever with a single tab, so the blocks are stacked in reading order and the
    formulas use ';' as argument separator (Italian locale).
    """
    rows: list[list] = []
    trades = {"first": None, "last": None, "headers": []}

    for block in build_report(data):
        if block["kind"] == "heading":
            rows.append([block["title"]])
            if block.get("subtitle"):
                rows.append([block["subtitle"]])
        elif block["kind"] == "pairs":
            rows.append([block["title"]])
            rows.extend([list(r) for r in block["rows"]])
        else:
            rows.append([block["title"]])
            rows.append(list(block["headers"]))
            if block["rows"]:
                if block.get("is_trades"):
                    trades = {"first": len(rows) + 1, "last": len(rows) + len(block["rows"]),
                              "headers": block["headers"]}
                rows.extend([list(r) for r in block["rows"]])
                if block["totals"]:
                    rows.append(list(block["totals"]))
            else:
                rows.append(["Nessun record nel periodo"])
        rows.append([])

    ctx = {"sheet": "", "first": trades["first"], "last": trades["last"], "arg": ";",
           "column": lambda label: (col_letter(trades["headers"].index(label) + 1)
                                    if label in trades["headers"] else "")}
    resolved = []
    for number, row in enumerate(rows, start=1):
        row_ctx = dict(ctx, row=number)
        resolved.append([resolve(cell, row_ctx) if trades["first"] else
                         "" if isinstance(cell, str) and cell.startswith("{{") else cell
                         for cell in row])

    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as fh:
        csv.writer(fh, delimiter=";").writerows(
            [[fmt_it(c) for c in row] for row in resolved])
    return path


# --------------------------------------------------------------------------- #
# XLSX output
# --------------------------------------------------------------------------- #

EUR_FMT = '#,##0.00\u00a0"€";[Red]-#,##0.00\u00a0"€"'
PRICE_FMT = '#,##0.00####'
PCT_FMT = '0.00%;[Red]-0.00%'
DATE_FMT = 'dd/mm/yyyy hh:mm:ss'

MONEY_COLS = {"Swap", "Commissioni", "Lordo EUR", "Netto EUR", "Saldo EUR", "Importo EUR",
              "Netto medio", "Saldo fine giornata"}
PRICE_COLS = {"Prezzo entrata", "Prezzo uscita", "Prezzo ordine", "Stop loss", "Take profit"}
PCT_LABELS = {"P/L % sul saldo", "Win rate"}
SHEETS = [("summary", "Riepilogo"), ("trades", "Operazioni"), ("daily", "Andamento"),
          ("orders", "Ordini"), ("positions", "Posizioni"), ("transactions", "Transazioni")]


def number_format(label, value):
    if isinstance(value, dt.datetime):
        return DATE_FMT
    if isinstance(value, dt.date):
        return "dd/mm/yyyy"
    if label in MONEY_COLS:
        return EUR_FMT
    if label in PRICE_COLS:
        return PRICE_FMT
    if label in PCT_LABELS:
        return PCT_FMT
    if label == "Tasso conversione":
        return "0.0000000000"
    if label in ("Durata (min)", "Punti", "Profit factor"):
        return "0.00"
    if label in ("Operazioni", "Vinte", "Perse", "Strumenti negoziati"):
        return "0"
    return "0.####" if isinstance(value, float) else "General"


# Diverging pair for gain/loss, validated against both surfaces with the palette
# checker: green/red reads as a single colour under deuteranopia (dE 4.1 on white,
# against a floor of 8), blue/red clears every check (dE 23.8).
CHART_PROFIT, CHART_LOSS = "2A78D6", "D03B3B"
CHART_INK, CHART_GRID = "52514E", "E1E0D9"
# One hue, light to dark: duration bands are ordered magnitude, not identity.
BLUE_RAMP = ["86B6EF", "6DA7EC", "5598E7", "3987E5", "2A78D6", "256ABF", "1C5CAB", "184F95"]


def style_chart(chart, title: str, legend: bool = False):
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties

    chart.title = title
    chart.height, chart.width = 7.5, 14
    if not legend:
        chart.legend = None
    elif chart.legend:
        chart.legend.position = "b"
    chart.x_axis.majorGridlines = None
    hairline = GraphicalProperties(ln=LineProperties(solidFill=CHART_GRID))
    chart.x_axis.spPr, chart.y_axis.spPr = hairline, hairline
    return chart


def paint_by_sign(series, values):
    """Colour each bar from the sign of its own value."""
    from openpyxl.chart.marker import DataPoint
    from openpyxl.chart.shapes import GraphicalProperties

    series.data_points = [
        DataPoint(idx=i, spPr=GraphicalProperties(
            solidFill=CHART_PROFIT if (value or 0) >= 0 else CHART_LOSS))
        for i, value in enumerate(values)]


def add_charts(wb, placed: dict, data: dict) -> None:
    """Four charts, each anchored beside the block it reads."""
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.marker import DataPoint
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.utils import get_column_letter

    def block_ref(spot, label, header=False):
        column = spot["headers"].index(label) + 1
        return Reference(wb[spot["sheet"]], min_col=column,
                         min_row=spot["first"] - (1 if header else 0), max_row=spot["last"])

    def categories(spot):
        return Reference(wb[spot["sheet"]], min_col=1,
                         min_row=spot["first"], max_row=spot["last"])

    symbols = placed.get("symbols")
    if symbols:
        rows = by_symbol(data)
        chart = BarChart()
        chart.type, chart.overlap = "bar", 100          # horizontal bars, one series
        chart.add_data(block_ref(symbols, "Netto EUR", header=True), titles_from_data=True)
        chart.set_categories(categories(symbols))
        paint_by_sign(chart.series[0], [row[4] for row in rows])
        style_chart(chart, "Netto per strumento (EUR)")
        wb[symbols["sheet"]].add_chart(chart, f"I{symbols['first'] - 1}")

        split = BarChart()
        split.type, split.grouping, split.overlap = "col", "stacked", 100
        for label, colour in (("Vinte", CHART_PROFIT), ("Perse", CHART_LOSS)):
            split.add_data(block_ref(symbols, label, header=True), titles_from_data=True)
            split.series[-1].graphicalProperties = GraphicalProperties(solidFill=colour)
        split.set_categories(categories(symbols))
        style_chart(split, "Operazioni vinte e perse", legend=True)
        wb[symbols["sheet"]].add_chart(split, f"I{symbols['first'] + 15}")

    durations = placed.get("durations")
    if durations:
        chart = BarChart()
        chart.type, chart.overlap = "col", 100
        chart.add_data(block_ref(durations, "Operazioni", header=True), titles_from_data=True)
        chart.set_categories(categories(durations))
        span = max(durations["last"] - durations["first"] + 1, 1)
        chart.series[0].data_points = [
            DataPoint(idx=i, spPr=GraphicalProperties(
                solidFill=BLUE_RAMP[min(int(i * len(BLUE_RAMP) / span), len(BLUE_RAMP) - 1)]))
            for i in range(span)]
        style_chart(chart, "Operazioni per durata")
        wb[durations["sheet"]].add_chart(chart, f"I{durations['first'] + 31}")

    daily = placed.get("daily")
    if daily and daily["last"] > daily["first"]:
        sheet = wb[daily["sheet"]]
        curve = LineChart()
        curve.add_data(block_ref(daily, "Saldo fine giornata", header=True),
                       titles_from_data=True)
        curve.set_categories(categories(daily))
        curve.series[0].graphicalProperties = GraphicalProperties()
        curve.series[0].graphicalProperties.line.solidFill = CHART_PROFIT
        curve.series[0].graphicalProperties.line.width = 22000
        curve.series[0].smooth = False
        style_chart(curve, "Curva del saldo (EUR)")
        sheet.add_chart(curve, f"{get_column_letter(len(daily['headers']) + 2)}{daily['first']}")

        bars = BarChart()
        bars.type, bars.overlap = "col", 100
        bars.add_data(block_ref(daily, "Netto EUR", header=True), titles_from_data=True)
        bars.set_categories(categories(daily))
        paint_by_sign(bars.series[0], [row[2] for row in by_day(data)])
        style_chart(bars, "Netto giornaliero (EUR)")
        sheet.add_chart(bars, f"{get_column_letter(len(daily['headers']) + 2)}{daily['first'] + 16}")


def write_xlsx(data: dict, path: str) -> str:
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    base = Font(name="Arial", size=10)
    bold = Font(name="Arial", size=10, bold=True)
    head_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F3864")
    title_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="2E5496")
    doc_font = Font(name="Arial", size=16, bold=True, color="1F3864")
    sub_font = Font(name="Arial", size=10, color="7F7F7F")
    win_font = Font(name="Arial", size=10, bold=True, color="1E7A46")
    loss_font = Font(name="Arial", size=10, bold=True, color="C00000")
    stripe = PatternFill("solid", fgColor="F2F5FA")
    total_fill = PatternFill("solid", fgColor="DCE3F0")
    band_fill = PatternFill("solid", fgColor="F7F9FC")
    thin = Side(style="thin", color="D9D9D9")
    box = Border(top=thin, bottom=thin, left=thin, right=thin)

    blocks = build_report(data)
    wb = Workbook()
    wb.remove(wb.active)
    widths: dict[str, dict[int, int]] = {}
    trades = {"sheet": None, "first": None, "last": None, "headers": []}
    placed: dict[str, dict] = {}       # where each table landed, for the charts
    pending: list[tuple] = []          # cells holding a token, resolved at the end

    for group, sheet_name in SHEETS:
        group_blocks = [b for b in blocks if b["group"] == group]
        if not group_blocks:
            continue
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        widths[sheet_name] = {}
        cursor = 0                     # last written row; openpyxl's max_row is 1 when empty

        def measure(col, text):
            widths[sheet_name][col] = max(widths[sheet_name].get(col, 10),
                                          min(len(str(text)) + 3, 30))

        for block in group_blocks:
            if block["kind"] == "heading":
                cursor += 1
                cell = ws.cell(row=cursor, column=1, value=block["title"])
                cell.font = doc_font
                ws.row_dimensions[cursor].height = 22
                if block.get("subtitle"):
                    cursor += 1
                    sub = ws.cell(row=cursor, column=1, value=block["subtitle"])
                    sub.font = sub_font
                    measure(1, block["subtitle"])
                cursor += 1
                continue

            cursor += 1
            title = ws.cell(row=cursor, column=1, value=block["title"])
            title.font, title.fill = title_font, title_fill
            span = len(block["headers"]) if block["kind"] == "table" else 2
            for c in range(2, span + 1):
                ws.cell(row=cursor, column=c).fill = title_fill
            ws.row_dimensions[cursor].height = 20

            if block["kind"] == "pairs":
                banded = block["title"] != "DATI DEL CONTO"
                for label, value in block["rows"]:
                    cursor += 1
                    r = cursor
                    key = ws.cell(row=r, column=1, value=label)
                    key.font = base
                    cell = ws.cell(row=r, column=2)
                    if banded:
                        key.fill = cell.fill = band_fill
                    if isinstance(value, str) and value.startswith("{{"):
                        pending.append((cell, value, label))
                    else:
                        cell.value = value
                        cell.number_format = number_format(label, value)
                    cell.font = base
                    cell.alignment = Alignment(horizontal="right")
                    measure(1, label)
                    measure(2, value)
                cursor += 1
                continue

            cursor += 1
            header_row = cursor
            for c, label in enumerate(block["headers"], start=1):
                cell = ws.cell(row=header_row, column=c, value=label)
                cell.font, cell.fill, cell.border = head_font, head_fill, box
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
                measure(c, label)
            ws.row_dimensions[header_row].height = 28

            if not block["rows"]:
                cursor += 1
                ws.cell(row=cursor, column=1,
                        value="Nessun record nel periodo").font = base
                cursor += 1
                continue

            first = header_row + 1
            for n, row in enumerate(block["rows"]):
                cursor += 1
                r = cursor
                for c, (label, value) in enumerate(zip(block["headers"], row), start=1):
                    cell = ws.cell(row=r, column=c)
                    cell.font, cell.border = base, box
                    if isinstance(value, str) and value.startswith("{{"):
                        pending.append((cell, value, label))
                    else:
                        cell.value = value
                        cell.number_format = number_format(label, value)
                    if n % 2:
                        cell.fill = stripe
                    if label == "Esito" and value in ("WIN", "LOSS"):
                        cell.font = win_font if value == "WIN" else loss_font
                    measure(c, fmt_intl(value))
            last = cursor

            spot = {"sheet": sheet_name, "first": first, "last": last,
                    "headers": block["headers"]}
            placed[block.get("id") or block["group"]] = spot
            if block.get("is_trades"):
                trades = spot
                ws.freeze_panes = ws.cell(row=first, column=1).coordinate

            # Where the money went, and which trades were left running: the same
            # diverging pair as the charts, so cell and bar say the same thing.
            column = lambda label: get_column_letter(block["headers"].index(label) + 1)
            for label in ("Netto EUR", "P/L % sul saldo"):
                if label in block["headers"]:
                    col = column(label)
                    ws.conditional_formatting.add(
                        f"{col}{first}:{col}{last}",
                        ColorScaleRule(start_type="min", start_color=CHART_LOSS,
                                       mid_type="num", mid_value=0, mid_color="FFFFFF",
                                       end_type="max", end_color=CHART_PROFIT))
            for label in ("Durata (min)", "Operazioni"):
                if label in block["headers"]:
                    col = column(label)
                    ws.conditional_formatting.add(
                        f"{col}{first}:{col}{last}",
                        DataBarRule(start_type="min", end_type="max", color="8EA9DB"))
            if block["totals"]:
                cursor += 1
                r = cursor
                for c, (label, value) in enumerate(zip(block["headers"], block["totals"]),
                                                   start=1):
                    cell = ws.cell(row=r, column=c)
                    cell.font, cell.fill, cell.border = bold, total_fill, box
                    if isinstance(value, str) and value.startswith("{{"):
                        pending.append((cell, value, label))
                        cell.number_format = EUR_FMT
                    elif value:
                        cell.value = value
            cursor += 1

    ctx = {"sheet": f"'{trades['sheet']}'!" if trades["sheet"] else "",
           "first": trades["first"], "last": trades["last"], "arg": ",",
           "column": lambda label: (get_column_letter(trades["headers"].index(label) + 1)
                                    if label in trades["headers"] else "")}
    for cell, token, label in pending:
        cell.value = (resolve(token, dict(ctx, row=cell.row)) if trades["first"] else None)
        # A formula cell has no value to infer a format from, so go by its column.
        cell.number_format = PCT_FMT if label in PCT_LABELS else number_format(label, 0.0)

    add_charts(wb, placed, data)

    for sheet_name, cols in widths.items():
        for col, width in cols.items():
            wb[sheet_name].column_dimensions[get_column_letter(col)].width = width

    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)
    shrink_xlsx(path)
    return path


def shrink_xlsx(path: str) -> int:
    """Repack the workbook as small as the format allows, and return its size.

    Uploading a workbook through an API that takes the bytes inline leaves very
    little room, so drop the parts nothing here relies on - the 10 KB colour
    theme (no style below refers to a theme colour) and the document properties
    - and recompress the rest at maximum deflate.
    """
    import re as _re
    import zipfile

    drop = ("xl/theme/theme1.xml", "docProps/app.xml", "docProps/core.xml")
    with zipfile.ZipFile(path) as zf:
        parts = {i.filename: zf.read(i.filename) for i in zf.infolist()}

    for name in drop:
        parts.pop(name, None)
    for name in ("[Content_Types].xml", "xl/_rels/workbook.xml.rels", "_rels/.rels"):
        if name in parts:
            xml = parts[name].decode("utf-8")
            xml = _re.sub(r"<(?:Override|Relationship)\b[^>]*?"
                          r"(?:theme/theme1\.xml|docProps/(?:app|core)\.xml)[^>]*?/>", "", xml)
            parts[name] = xml.encode("utf-8")

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED, compresslevel=9) as zf:
        for name, blob in parts.items():
            zf.writestr(name, blob)
    return os.path.getsize(path)


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("statement", help="cTrader statement .html file")
    ap.add_argument("--xlsx", help="write a Google Sheets ready workbook here")
    ap.add_argument("--csv-dir", help="write one CSV per section into this directory")
    ap.add_argument("--gsheet-csv", metavar="FILE",
                    help="write a single semicolon CSV with every section stacked, "
                         "ready to upload to Google Drive as a native Google Sheet")
    args = ap.parse_args(argv)

    if not args.xlsx and not args.csv_dir and not args.gsheet_csv:
        ap.error("choose at least one of --xlsx / --csv-dir / --gsheet-csv")

    data = parse_statement(args.statement)
    counts = ", ".join(f"{name}: {len(s['rows'])}" for name, s in data["sections"].items())
    print(f"parsed {args.statement} -> {counts}")
    for warning in check_consistency(data):
        print(f"ATTENZIONE: {warning}", file=sys.stderr)

    if args.csv_dir:
        for path in write_csvs(data, args.csv_dir):
            print("wrote", path)
    if args.gsheet_csv:
        print("wrote", write_gsheet_csv(data, args.gsheet_csv))
    if args.xlsx:
        print("wrote", write_xlsx(data, args.xlsx))
    return 0


if __name__ == "__main__":
    sys.exit(main())
